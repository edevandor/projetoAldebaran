"""Gera dados sinteticos de vendas no formato ERP hierarquico.

Cria um ficheiro XLSX em data/raw/ que segue a mesma estrutura dos
relatorios originais: blocos por responsavel e artigo, com linhas de
venda individuais. Usado para demonstracao e testes do pipeline sem
expor dados de clientes.

Uso:
    python scripts/gerar_dados_sinteticos.py
"""

import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment

# --- Configuracoes ---
PRODUTOS = [
    ("COMP-A", "Compressor A 7.5CV"),
    ("COMP-B", "Compressor B 10CV"),
    ("COMP-C", "Compressor C 15CV Turbo"),
    ("VALV-S", "Valvula Solenoide 1/2\""),
    ("VALV-R", "Valvula Retencao 3/4\""),
    ("FILT-O", "Filtro Oleo SP-200"),
    ("FILT-A", "Filtro Ar PA-150"),
    ("MOTO-E", "Motor Eletrico 5CV 4 polos"),
    ("MOTO-G", "Motor Eletrico 7.5CV 6 polos"),
    ("RESIS", "Resistencia Blindada 12kW"),
    ("PS-100", "Pressostato Diferencial PS-100"),
    ("TERM-K", "Termostato Eletronico TK-2000"),
    ("CABO-4", "Cabo PP 4mm² por metro"),
    ("CABO-6", "Cabo PP 6mm² por metro"),
    ("DUTO-F", "Duto Flexivel 100mm"),
]

RESPONSAVEIS = ["Carlos Silva", "Ana Oliveira", "Roberto Lima"]
ARTIGOS = ["Venda Direta", "Orcamento", "Reposicao"]

COL_PRODUTO = 0       # A
COL_UND = 5           # F
COL_QTD = 6           # G
COL_DATA = 7          # H
COL_PROTOCOLO = 8     # I
COL_NF = 9            # J
COL_TOTAL = 13        # N


def _gerar_nf(idx_venda: int) -> str:
    """Gera numero de NF no formato '001-NNNNNN'."""
    return f"001-{100000 + idx_venda:06d}"


def _gerar_protocolo() -> str:
    """Gera protocolo de 6 digitos."""
    import random
    return f"{random.randint(100000, 999999)}"


def gerar_xlsx(destino: Path, num_vendas: int = 50):
    """Gera um XLSX sintetico com estrutura ERP hierarquica."""
    import random
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MapaVendas"

    bold_font = Font(bold=True)

    linha_atual = 1

    for resp_idx, responsavel in enumerate(RESPONSAVEIS):
        vendas_por_artigo = num_vendas // len(ARTIGOS)

        for artigo in ARTIGOS:
            # --- Linha de responsavel ---
            ws.cell(row=linha_atual, column=COL_PRODUTO + 1,
                    value=f"Responsavel: {responsavel}")
            ws.cell(row=linha_atual, column=COL_PRODUTO + 1).font = bold_font
            linha_atual += 1

            # --- Linha de artigo ---
            ws.cell(row=linha_atual, column=COL_PRODUTO + 1,
                    value=f"Artigo: {artigo}")
            ws.cell(row=linha_atual, column=COL_PRODUTO + 1).font = bold_font
            linha_atual += 1

            # --- Cabecalho de colunas ---
            cabecalhos = [
                "Produto", "Estoque", "", "", "", "Und", "Qtde", "Data Emissao",
                "Protocolo", "NF Nº", "", "", "", "Total Venda",
            ]
            for i, cab in enumerate(cabecalhos):
                cell = ws.cell(row=linha_atual, column=i + 1, value=cab)
                cell.font = bold_font
            linha_atual += 1

            # --- Linhas de venda ---
            base_data = datetime.date(2025, 1, 1)
            for i in range(min(vendas_por_artigo, 20)):
                produto = random.choice(PRODUTOS)
                codigo, descricao = produto
                qtd = round(random.uniform(1, 50), 2)
                preco_unit = round(random.uniform(50, 5000), 2)
                total = round(qtd * preco_unit, 2)
                dias = random.randint(0, 540)
                data = base_data + datetime.timedelta(days=dias)
                nf = _gerar_nf(resp_idx * 100 + i)
                protocolo = _gerar_protocolo()
                und = random.choice(["UN", "PC", "M", "KG"])

                ws.cell(row=linha_atual, column=COL_PRODUTO + 1,
                        value=f"{codigo} - {descricao}")
                ws.cell(row=linha_atual, column=COL_UND + 1, value=und)
                ws.cell(row=linha_atual, column=COL_QTD + 1, value=qtd)
                ws.cell(row=linha_atual, column=COL_DATA + 1, value=data)
                ws.cell(row=linha_atual, column=COL_PROTOCOLO + 1, value=protocolo)
                ws.cell(row=linha_atual, column=COL_NF + 1, value=nf)
                ws.cell(row=linha_atual, column=COL_TOTAL + 1, value=total)

                linha_atual += 1

            # --- Linha de total do artigo ---
            ws.cell(row=linha_atual, column=COL_PRODUTO + 1,
                    value=f"Total: {artigo}")
            ws.cell(row=linha_atual, column=COL_PRODUTO + 1).font = bold_font
            linha_atual += 1

        # --- Linha de total do responsavel ---
        ws.cell(row=linha_atual, column=COL_PRODUTO + 1,
                value=f"Total: Responsavel: {responsavel}")
        ws.cell(row=linha_atual, column=COL_PRODUTO + 1).font = bold_font
        linha_atual += 1

    # Salvar
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    print(f"Dados sinteticos gerados: {destino.resolve()}")
    print(f"  Responsaveis: {len(RESPONSAVEIS)}")
    print(f"  Artigos: {len(ARTIGOS)}")
    print(f"  Total linhas geradas: ~{linha_atual}")
    return destino


if __name__ == "__main__":
    destino = Path("data/raw/vendas_sinteticas.xlsx")
    gerar_xlsx(destino, num_vendas=50)
    print("Execute o pipeline com: python -c 'from src.pipeline import run_pipeline; print(run_pipeline())'")
