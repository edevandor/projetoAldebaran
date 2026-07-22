"""Parser de XLSX de mapas estatísticos de vendas (formato ERP hierárquico).

Lê ficheiros com estrutura de blocos por artigo, extrai apenas as linhas
de venda e mapeia para o modelo de dados do pipeline.
"""

import datetime
from enum import Enum, auto
from pathlib import Path
from typing import Any, List

import openpyxl

# --- Constantes de mapeamento de colunas (0-indexed) ---
COL_PRODUTO = 0          # A
COL_ESTOQUE = 3          # D (ignorado)
COL_UNIDADE = 5          # F
COL_QUANTIDADE = 6       # G
COL_DATA_EMISSAO = 7     # H
COL_PROTOCOLO = 8        # I
COL_NF_NUMERO = 9        # J
COL_TOTAL_VENDA = 13     # N
_MAX_COL_INDEX = 15      # Máximo índice para checagem de linha vazia


class _Estado(Enum):
    """Estados da máquina de estados do parser."""
    METADATA = auto()
    ARTICLE_HEADER = auto()
    COLUMN_HEADER = auto()
    DATA = auto()
    TOTAL = auto()


def _cell_str(val: Any) -> str:
    """Converte célula para string, tratando None como string vazia."""
    if val is None:
        return ""
    return str(val).strip()


def _cell_num(val: Any) -> float:
    """Converte célula para float, tratando None como 0."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _cell_date(val: Any) -> datetime.date | None:
    """Converte célula para date, retorna None se inválido."""
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


def _extrair_tipo_documento(numero_documento: str) -> str:
    """Extrai o prefixo (tipo) do número do documento.

    Ex: '001-100001' → '001'
    """
    if not numero_documento:
        return ""
    return numero_documento.split("-", 1)[0]


def _normalize_prefix(texto: str) -> str:
    """Remove acentos para comparação de prefixos.

    Ex: 'Responsável:' → 'Responsavel:'
    """
    substituicoes = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e", "í": "i", "ó": "o",
        "ô": "o", "õ": "o", "ú": "u", "ç": "c",
        "Á": "A", "À": "A", "Ã": "A", "Â": "A",
        "É": "E", "Ê": "E", "Í": "I", "Ó": "O",
        "Ô": "O", "Õ": "O", "Ú": "U", "Ç": "C",
    }
    for acentuada, sem in substituicoes.items():
        texto = texto.replace(acentuada, sem)
    return texto


def _linha_eh_responsavel(texto: str) -> bool:
    """Linha começa com 'Responsável:' (com ou sem acento, case-insensitive)."""
    return _normalize_prefix(texto).lower().startswith("responsavel:")


def _linha_eh_total_responsavel(texto: str) -> bool:
    """Linha começa com 'Total: Responsável:' (com ou sem acento, case-insensitive)."""
    return _normalize_prefix(texto).lower().startswith("total: responsavel:")


def _parse_responsavel(texto: str) -> str:
    """Extrai o nome do responsável de 'Responsável: NOME'.

    Aceita com ou sem acento, case-insensitive.
    Retorna string vazia se não conseguir.
    """
    if not texto:
        return ""
    normalizado = _normalize_prefix(texto).lower()
    prefixo = "responsavel:"
    idx = normalizado.find(prefixo)
    if idx >= 0:
        nome = texto[idx + len(prefixo):].strip()
        return nome
    return ""


def _parse_artigo(texto: str) -> str:
    """Extrai o nome do artigo de 'Artigo: NOME'.

    Retorna string vazia se não conseguir extrair.
    """
    if not texto:
        return ""
    prefixo = "Artigo:"
    if texto.startswith(prefixo):
        return texto[len(prefixo):].strip()
    return ""


def _extrair_campos(row: tuple) -> dict | None:
    """Extrai os campos de uma linha de dados.

    Retorna None apenas se o produto estiver vazio (linha sem dados).
    Datas ausentes são preservadas como None — a validação decide.
    """
    produto = _cell_str(row[COL_PRODUTO])
    if not produto:
        return None

    data_raw = row[COL_DATA_EMISSAO] if len(row) > COL_DATA_EMISSAO else None
    data_emissao = _cell_date(data_raw)
    # data_emissao pode ser None — não descartar a linha.
    # Migração de servidor em 2023 gerou falhas de datação.
    # A validação ou o usuário decide o que fazer com datas ausentes.

    numero_doc = _cell_str(row[COL_NF_NUMERO] if len(row) > COL_NF_NUMERO else None)

    return {
        "produto": produto,
        "data_emissao": data_emissao,
        "tipo_documento": _extrair_tipo_documento(numero_doc),
        "numero_documento": numero_doc,
        "protocolo": _cell_str(row[COL_PROTOCOLO] if len(row) > COL_PROTOCOLO else None),
        "valor_total_venda": _cell_num(row[COL_TOTAL_VENDA] if len(row) > COL_TOTAL_VENDA else None),
        "quantidade": _cell_num(row[COL_QUANTIDADE] if len(row) > COL_QUANTIDADE else None),
        "unidade_medida": _cell_str(row[COL_UNIDADE] if len(row) > COL_UNIDADE else None),
    }


def parse_xlsx(path: str | Path) -> List[dict]:
    """Parseia um XLSX de mapa estatístico de vendas.

    Retorna uma lista de dicionários, um por linha de venda,
    com as chaves do modelo de dados do pipeline.
    Cada linha inclui 'responsavel' (vendedor/representante)
    e 'artigo' (categoria de produto) herdados da hierarquia.
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"Ficheiro sem worksheet activa: {path}")

    resultados: List[dict] = []
    estado = _Estado.METADATA
    current_responsavel: str = ""
    current_artigo: str = ""

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cell_1 = _cell_str(row[COL_PRODUTO])

        # --- METADATA ---
        if estado == _Estado.METADATA:
            if _linha_eh_responsavel(cell_1):
                current_responsavel = _parse_responsavel(cell_1)
                continue
            if cell_1.startswith("Artigo:"):
                current_artigo = _parse_artigo(cell_1)
                estado = _Estado.ARTICLE_HEADER
                continue
            if _linha_eh_total_responsavel(cell_1):
                continue
            continue

        # --- ARTICLE_HEADER ---
        if estado == _Estado.ARTICLE_HEADER:
            estado = _Estado.COLUMN_HEADER
            continue

        # --- COLUMN_HEADER (falls through to DATA — mesma row) ---
        if estado == _Estado.COLUMN_HEADER:
            estado = _Estado.DATA

        # --- DATA ---
        if estado == _Estado.DATA:
            if _linha_eh_total_responsavel(cell_1):
                estado = _Estado.TOTAL
                continue
            if cell_1.startswith("Total:"):
                estado = _Estado.TOTAL
                continue
            if cell_1.startswith("Artigo:"):
                current_artigo = _parse_artigo(cell_1)
                estado = _Estado.ARTICLE_HEADER
                continue
            if _linha_eh_responsavel(cell_1):
                current_responsavel = _parse_responsavel(cell_1)
                continue
            if cell_1 == "Produto":
                continue
            if all(c is None for c in row[:_MAX_COL_INDEX]):
                continue

            campos = _extrair_campos(row)
            if campos is not None:
                campos["responsavel"] = current_responsavel
                campos["artigo"] = current_artigo
                resultados.append(campos)
            continue

        # --- TOTAL ---
        if estado == _Estado.TOTAL:
            if _linha_eh_responsavel(cell_1):
                current_responsavel = _parse_responsavel(cell_1)
                continue
            if cell_1.startswith("Artigo:"):
                current_artigo = _parse_artigo(cell_1)
                estado = _Estado.ARTICLE_HEADER
                continue
            if _linha_eh_total_responsavel(cell_1):
                continue
            continue

    wb.close()
    return resultados
